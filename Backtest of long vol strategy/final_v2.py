from pandas import read_csv,read_excel,cut,ExcelWriter,value_counts,bdate_range,concat
from smtplib  import SMTP   #傳送郵件
from zipfile  import ZipFile
from datetime import datetime
from dateutil.relativedelta import relativedelta
from datetime import timedelta
from email.mime.text import MIMEText               #專門傳送正文
from email.mime.multipart import MIMEMultipart     #傳送多個部分
from email.mime.application import MIMEApplication #傳送附件
from openpyxl import load_workbook
#%%
def time_for_calculate(time_list):
    time_fix=[]
    for _,time in enumerate(time_list):
        if len(str(time))==1:
            time_fix.append(int('24000'+str(time)))
        elif len(str(time))==2:
            time_fix.append(int('2400'+str(time)))
        elif len(str(time))==3:
            time_fix.append(int('240'+str(time)))
        elif len(str(time))==4:
            time_fix.append(int('24'+str(time)))
        elif len(str(time))==5:
            time_fix.append(int(   str(24+int( str(time)[0] ))+str(time)[1::]   ))
        else:
            time_fix.append(time)
    return time_fix

def datetime_range(start, end, delta):
    current = start
    while current < end:
        yield current
        current += delta
def date_process(date):
    file_name= datetime.strftime(date, "Daily_%Y_%m_%d")
    date_index=datetime.strftime(date, "%Y%m%d")
    Expiration_Date=bdate_range(start=str(date.year), end=str(date.year+1), freq='WOM-3WED')
    Nearby_future=datetime.strftime(date, "%Y%m")
    Back_future=datetime.strftime(date+ relativedelta(months=1), "%Y%m")
    return file_name,date_index,Expiration_Date,Nearby_future,Back_future
def dts_int(work_time):
    if work_time==0:     
        dts_index = [ dt.strftime('%H%M%S')
                for dt in datetime_range(datetime(2016,9,1,15), datetime(2016, 9, 2,5),timedelta(minutes=30))]    
        dts = [ dt.strftime('%H%M%S')
                if dt<datetime(2016,9,2,0) else str(24+int(dt.strftime('%H%M%S')[:2]))+ dt.strftime('%H%M%S')[2::]
                for dt in datetime_range(datetime(2016,9,1,15,30), datetime(2016, 9, 2,5),timedelta(minutes=30))]#跨天處理
        dts.insert(0,'144500');dts.append('291500')
        dts_int=[int(num) for num in dts]
        return dts_int,dts_index
    if work_time==1:#日盤
        dts = [dt.strftime('%H%M%S') for dt in datetime_range(datetime(2016, 9, 1, 9), datetime(2016, 9, 1,14),timedelta(minutes=30))]
        dts.append('134500');dts.insert(0,'084400')
        dts_int=[int(num) for num in dts]
        return dts_int
#%%時間序列處理
date=datetime.today()
file_name,date_index,Expiration_Date,Nearby_future,Back_future=date_process(date)

#%%資料解壓縮
zip_name =r'T:\期貨交易所每筆成交資料\期貨'+'\\'+str(date.year)+'\\'+file_name+'.zip'
file_dir =r'T:\Future_statistics\daily' #解壓後的文件放在該目錄下
with ZipFile(zip_name, 'r') as myzip:
    for file in myzip.namelist():
        myzip.extract(file,file_dir)       
#%%時間index
dts_int_n,dts_index_n=dts_int(work_time=0)#夜盤
dts_int_w=dts_int(work_time=1)#日盤
 #%%
# =============================================================================
# 當日點數計算
# =============================================================================
for work_time in [0,1]:#夜盤、日盤
    a_list=[1,2,3,4,5,6,7,8,9,10]
    output_final=[];time_list=[];temporary_list=[]
    for _,a in enumerate(a_list):
        df = read_csv(file_dir+'\\'+file_name+'.rpt', encoding = 'Big5',low_memory=False)
        df =df.loc[df['商品代號']=='TX     ']
        if work_time==0:#夜盤
            df=df.loc[( (df['成交日期']!=int(date_index)) & (df['成交時間']<=235959) ) | ((df['成交日期']==int(date_index)) & (df['成交時間']<=50000))]
        elif work_time==1:
            df =df.loc[(70000<df['成交時間']) & (df['成交時間']<145000)]

        if date< Expiration_Date[date.month-1]:
            df=df[df['到期月份(週別)']==Nearby_future+'     ']
        elif date>= Expiration_Date[date.month-1]:
            df=df[df['到期月份(週別)']==Back_future+'     ']
        x=df.iloc[:,4]
        time=df.iloc[:,3]
        x.reset_index(inplace=True,drop=True)
        time.reset_index(inplace=True,drop=True)
        temp=x[0];time_temp=[]
        for i in range(x.shape[0]):
            if abs(x[i] - temp) >= a:
                time_temp.append(time[i])
                temp = x[i]
        if work_time==0:#夜盤
            time_list.append( time_for_calculate(time_temp) )
        elif work_time==1:
            time_list.append(time_temp.copy())
    #%%當日點數時間計算
    for point in range(len(time_list)):
        if work_time==0:
            se1=cut(time_list[point],dts_int_n)  #分時間
        elif work_time==1:
            se1=cut(time_list[point],dts_int_w)  #分時間
        temporary=value_counts(se1)        #分時間計數
        temporary.sort_index(inplace=True)    #排序
        temporary_list.append(temporary)
    #格式處理
    only_one=concat(temporary_list,axis=1)
    only_one.columns=a_list
    if work_time==1:
        only_one.index=only_one.index.tolist()
        only_one['index']=dts_int_w[0:-1]
    elif work_time==0:#夜盤
        only_one['index']=dts_index_n
    only_one = only_one.set_index('index')
    only_one.loc['total']=only_one.sum()
    #輸出
    if work_time==1:
        path1 = r'T:\Future_statistics\點數記錄檔\日盤' +'\\'+file_name[6::]+'早盤統計.xlsx'
        writer = ExcelWriter(path1, engine = 'openpyxl')

    elif work_time==0:
        path2 = r'T:\Future_statistics\點數記錄檔\夜盤' +'\\'+file_name[6::]+'夜盤統計.xlsx'
        writer = ExcelWriter(path2, engine = 'openpyxl')
    only_one.to_excel(writer,sheet_name=file_name[6::])
    writer.save()
    
print("excel輸出完成！")
#%%
wb1= load_workbook(path1)
wb1_sheet = wb1.active
max_row=wb1_sheet.max_row
name=wb1_sheet.title[-5::]
for row in wb1_sheet.iter_rows(max_row=max_row, min_row=max_row,values_only=True):
    total=list(row)
total[0]=name
path_s=r"T:\Future_statistics\點數記錄檔\日盤統計檔.xlsx"
wb2= load_workbook(path_s)
wb2_sheet = wb2.active
wb2_sheet.append(total)
wb2.save(path_s)
print('統計檔修改完成！')
#%%
# =============================================================================
# Email
# =============================================================================
df_mail=read_excel(r'T:\Future_statistics\final_code\Email_Future.xls')
Email_list=df_mail['Email'].values.tolist()
Email_str=','.join(Email_list)


send_user = '新金融商品處<ewarrant@jsun.com>'   #發件人
receive_users = Email_str   #收件人，可為list
subject = date_index+'台指期點數統計'  #郵件主題
email_text = '執行檔路徑: T:\Future_statistics\\final_code\\final.exe\r\n\r\n\
資料路徑:\r\n\
    日盤: T:\Future_statistics\點數記錄檔\日盤\"日期"+早盤統計.xlsm\r\n\
    夜盤: T:\Future_statistics\點數記錄檔\夜盤\"日期"+夜盤統計.xlsm\r\n\
    日盤統計檔:T:\Future_statistics\點數記錄檔\日盤統計檔.xlsm\r\n\r\n\
寄件者設定: T:\Future_statistics\\final_code\Email_Future.xls\r\n\
2020/05/05起以固定delta reblance計算點數' #郵件正文

server_address = 'mx00.jsun.com'   #伺服器地址
mail_type = '1'    #郵件型別

#構造一個郵件體：正文 附件
msg = MIMEMultipart()
msg['Subject']=subject    #主題
msg['From']=send_user      #發件人
msg['To']=receive_users           #收件人

#構建正文
part_text=MIMEText(email_text)
msg.attach(part_text)             #把正文加到郵件體裡面去

#構建郵件附件
#附件1
part_attach1 = MIMEApplication(open(path1,'rb').read())   #開啟附件
part_attach1.add_header('Content-Disposition','attachment',filename=path1[-19::]) #為附件命名
msg.attach(part_attach1)   #新增附件
#附件2
part_attach2 = MIMEApplication(open(path2,'rb').read())   #開啟附件
part_attach2.add_header('Content-Disposition','attachment',filename=path2[-19::]) #為附件命名
msg.attach(part_attach2)   #新增附件
part_attach3 = MIMEApplication(open(path_s,'rb').read())   #開啟附件
part_attach3.add_header('Content-Disposition','attachment',filename=path_s[-10::]) #為附件命名
msg.attach(part_attach3)   #新增附件
# 傳送郵件 SMTP
smtp= SMTP(server_address,25)  # 連線伺服器，SMTP_SSL是安全傳輸

#smtp.login(send_user, password)
smtp.sendmail(send_user, Email_list, msg.as_string())  # 傳送郵件
print('郵件傳送成功！')
import os
path=r"T:\Future_statistics\daily"
for file in os.listdir(path):
    os.remove(path+"\\"+file)
print('刪除daily')  
    
    

    