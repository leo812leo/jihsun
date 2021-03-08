import pandas as pd
from sys import path
path.append(r'C:\Users\F128537908\Desktop\WCFADOX_Python')
from WCFAdox import PCAX
from datetime import datetime
from pandas.tseries.offsets import BDay 
import re
import numpy as np
from openpyxl import load_workbook
import xlwings as xw 
from email.mime.text import MIMEText               #專門傳送正文
from smtplib  import SMTP   #傳送郵件
import smtplib
from itertools import chain



def error(errortype,date_str):
    df_mail=pd.read_excel(r'Email_發行評等.xls')
    Email_list=df_mail['Email'].dropna().values.tolist()
    Email_str=','.join(Email_list)
    send_user = '新金融商品處<ewarrant@jsun.com>'   #發件人
    receive_users = Email_str   #收件人，可為list
    subject = "(Error)"+date_str+'發行人評等'  #郵件主題
    text="錯誤訊息:"+errortype
    #郵件正文
    server_address = 'mx00.jsun.com'   #伺服器地址
    #構造一個郵件體：正文 附件
    #構建正文
    msg = MIMEText(text,'plain','utf-8')
    #把正文加到郵件體裡面去
    msg['Subject']=subject    #主題
    msg['From']=send_user      #發件人
    msg['To']=receive_users           #收件人
    try:
        # 傳送郵件 SMTP
        smtp= SMTP(server_address,25)  # 連線伺服器，SMTP_SSL是安全傳輸
        # #smtp.login(send_user, password)
        smtp.sendmail(send_user, Email_list, msg.as_string())  # 傳送郵件
    except smtplib.SMTPException:
        print( "Error: 無法傳送郵件")
#---------------------------------------------------------------#
#檔案讀取
date=datetime.today()
date_str=datetime.strftime(date, "%Y%m%d")
# holiday=[Taiwan().holidays(date.year)[i][0] for i in range(len(Taiwan().holidays(date.year)))]
# holiday.append(datetime(2020,5,1).date())
df=pd.read_csv('https://www.twse.com.tw/holidaySchedule/holidaySchedule?response=csv&queryYear='+\
               str(datetime.now().year-1911),header=1,encoding='big5')
tem=[re.findall(r'(\d+月\d+日)',i.values[0])for _,i in df[['日期']].iterrows()]
holiday=[(datetime.strptime(str(datetime.now().year)+i, "%Y%m月%d日")).date() for i in chain(*tem)]



last_workday=datetime.now()- BDay(1)                            #看前一個工作天
last_workday=last_workday.date()
if last_workday in  holiday:                                    #遇國定假日再往前一天
    last_workday=last_workday- BDay(1) 

last2_workday=last_workday- BDay(1)                             #看前兩個工作天(權證庫存用)
last2_workday=last2_workday.date()
if last2_workday in  holiday:                                   #遇國定假日再往前一天
    last2_workday=last2_workday- BDay(1) 
    
file_name=datetime.strftime(last_workday, "%Y%m%d")+'.txt'

#TWSE+OTC
file_list=[]
for platform in ['WRTWSE','WROTC']:
    try:
        file=open('T:\權證資訊揭露平台每日價量1'+'\\'+ platform + file_name ,'r') 
        file_list.extend(file.readlines()) 
    except:
        error("讀取資料錯誤",date_str)                 
        input()
        
title='資料期間	取樣期間	權證代號	權證名稱	到期日	存續期間	執行價	價內程度	行使比例	權證收盤價	標的收盤價	委買隱波	委賣隱波	一段時間委買隱波	一段時間委賣隱波	一段時間委買隱波最大變動值	一段時間委賣隱波最大變動值	一段時間委託買賣波動率差值	買賣價差檔數	買賣價差比	權證委託買量	權證委託賣量	收盤隱波	委買報價波動率 \
	買賣權	最後更新日期	標的代號	標的名稱	券商代號	券商名稱	買1量*買1價之平均值	一段時間委買隱波	一段時間委賣隱波'
title   = title.split('\t')

#%%
# =============================================================================
# 1.資料匯入處理
# =============================================================================
out=[]
for num_of_warrant in range(len(file_list)):
    result=re.search(r'(^\d{8})(\d{3})(\d{5}[0-9P])(.*[購售]\d{2})\s+(\d{8})(.{4})(.{8})(.{9})(.{9})(.{7})(.{8})(.{8})(.{8})(.{8})(.{8})(.{8})(.{8})(.{9})(.{3})(.{8})(.{8})(.{8})(.{8})(.{9})([CP])(\d{8})(\d{1,5}[0-9RUL])\s+(\D+.*[\D1-3])[\s\D]+(\d.\d{2})\b(\D+)\s+(.{8})(.{8})(.{8})', file_list[num_of_warrant])
    out.append(result.groups())
df=pd.DataFrame(out,columns=title) #轉成DataFrame

#轉日期
to_datetime_col=['資料期間','到期日','最後更新日期']
df[to_datetime_col]=df[to_datetime_col].applymap(lambda s: datetime.strptime(s, "%Y%m%d"))
#轉數值
to_numeries_col=[1,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,30,31,32]
df.iloc[:,to_numeries_col]=df.iloc[:,to_numeries_col].apply(lambda s:s.str.replace('^0+', ''))
df.iloc[:,to_numeries_col]=df.iloc[:,to_numeries_col].apply(lambda s:s.str.replace('^\.', '0.'))
df.iloc[:,to_numeries_col]=df.iloc[:,to_numeries_col].apply(lambda s: pd.to_numeric(s,errors='coerce'))
#轉文字
to_str_col=['標的代號','標的名稱','券商代號','券商名稱']
df.loc[:,to_str_col]=df.loc[:,to_str_col].apply(lambda s: s.str.strip())
#index為權證代碼
df=df.set_index('權證代號')

del date,file,file_list,file_name,title,holiday
del platform,result,out,num_of_warrant
del to_datetime_col,to_numeries_col,to_str_col
 
#%%
# =============================================================================
# 2.前處理
'''
1.交易日期: 計算16個交易日為多少天日曆日
2.權證標的數計算
3.篩選 (1)超過15個工作天 (2)價外20%~價內10%
4.權證十檔以上之標的(conditionX)
'''
# =============================================================================

def append_warrant_underlying_num(dataframe):
    warrant_underlying_num=dataframe['標的代號'].value_counts()
    dataframe.loc[:,"同標的權證檔數"]=warrant_underlying_num[dataframe['標的代號'].values].values
    return dataframe

#計算剩餘16個交易日 為幾天日曆日
交易日表=pd.read_html("http://172.16.10.13/hedge/tradingday.php",encoding='big5')[4]
交易日表=交易日表.drop(交易日表.columns[0],axis=1)
交易日表.columns=交易日表.iloc[0,:]
交易日表=交易日表.drop(0,axis=0)
交易日表.index=交易日表.index.astype(int)
交易日表['交易日期']=pd.to_datetime(交易日表['交易日期'])
Calendar_days=(交易日表.iloc[16,0]-交易日表.iloc[0,0]).days

#交易日剩餘大於16天的資料#價外20%~價內10%
condition1=df["存續期間"]>=Calendar_days               
condition2=(df["價內程度"]>=-0.1) & (df["價內程度"]<=0.2) 
df=df[condition1 & condition2]

#篩選 權證十檔以上之標的
df=append_warrant_underlying_num(df)
conditionX=df["同標的權證檔數"]>=10
final_data=df[conditionX]#篩選結果1

del 交易日表,Calendar_days
del condition1,condition2,conditionX
#%%
# =============================================================================
# 3.(1)價格穩定分數 (2)價差大小分數
'''
1.分數計算    : (1)價格穩定分數 (2)價差大小分數
2.權證庫存    : 匯入前一日庫存資料
'''
# =============================================================================
#分數計算
final_data.loc[:,'價格穩定分數']=final_data.groupby('標的代號')['一段時間委買隱波最大變動值'].rank(pct=True,ascending=False,method='max')
final_data.loc[:,'價差大小分數']=final_data.groupby('標的代號')['買賣價差比'].rank(pct=True,ascending=False,method='max')


#OI
last_work_day_str=datetime.strftime(last_workday,"%Y%m%d")#日期
last2_work_day_str=datetime.strftime(last2_workday,"%Y%m%d")#日期
PX=PCAX("192.168.65.11");col="日期,代號,名稱,自營商庫存"
try:
    warrant_OI_data=PX.Pal_Data("權證日自營商進出排行","D",last2_work_day_str,last2_work_day_str,colist=col)
    warrant_OI_data.loc[:,'自營商庫存']=warrant_OI_data.loc[:,'自營商庫存'].astype(int)#轉數值
    warrant_OI_data=warrant_OI_data.set_index('代號')                     #轉INDEX
except: 
    error("cmoney讀取問題")
del col,last_workday,last2_work_day_str,last2_workday
#%%
# =============================================================================
# 4.(3)報價數量分數
'''
  1. 調整買1量*買1價最大值為(50萬)
  2. 篩選 權證前一日流通在外大於0的權證(condition3)
  3. 有效權證檔數>10                  (conditionX)
  4. 分數計算 (3)報價數量分數
  5. 總分計算依券商
'''
# =============================================================================

#篩選 在外大於0的權證
condition3=warrant_OI_data[warrant_OI_data['自營商庫存']<0].index.to_list() #oi與data不同表用list處理
filter_Intersection = df.index.isin(condition3)                            #篩選資料取交集
data_for_BidVolume=df[filter_Intersection]

#有效權證檔數>10
data_for_BidVolume=append_warrant_underlying_num(data_for_BidVolume)
conditionX=data_for_BidVolume["同標的權證檔數"]>=10    
data_for_BidVolume=data_for_BidVolume[conditionX]
data_for_BidVolume=data_for_BidVolume.loc[:,['標的代號','買1量*買1價之平均值']]                       #取要用的資料

#調整買1量*買1價最大值為(50萬)
data_for_BidVolume.loc[:,'買1量*買1價之平均值']=data_for_BidVolume.loc[:,'買1量*買1價之平均值'].apply(lambda s: 500 if s>500 else s)
#報價數量分數 
data_for_BidVolume['報價數量分數']=data_for_BidVolume.groupby('標的代號')['買1量*買1價之平均值'].rank(pct=True,method='max')
final_data.loc[data_for_BidVolume.index,'報價數量分數']=data_for_BidVolume['報價數量分數']
del condition3,conditionX,filter_Intersection,warrant_OI_data
#%%
# =============================================================================
# 5.券商總分計算
'''
  1. 總分計算 (1)依券商
'''
# =============================================================================
#評等表
def table_for_LetterGrade(score):
    if score>=70:
        Letter_Grade='A'
    elif score>=50:
        Letter_Grade='B'
    elif score>=20:
        Letter_Grade='C'
    else:
        Letter_Grade='D'
    return Letter_Grade

BrokerScore_Table=pd.pivot_table(final_data,index=['券商名稱'],values=['價格穩定分數','價差大小分數','報價數量分數'],aggfunc=[np.mean])
BrokerScore_Table.columns=BrokerScore_Table.columns.droplevel(0) #除掉多餘COLUMNS
BrokerScore_Table=BrokerScore_Table*100
BrokerScore_Table=(BrokerScore_Table.rank(pct=True,method='max')*100)         #再取一次PCT並化簡
BrokerScore_Table['總分']=BrokerScore_Table.mean(axis=1)         #算出總分並化簡
BrokerScore_Table=BrokerScore_Table.round(1)                     #每一檔權證權重相同 
BrokerScore_Table['評等']=BrokerScore_Table['總分'].map(table_for_LetterGrade)

#%%
# =============================================================================
# 6.交易員分數
'''
  1. 交易員資料讀取
  2. 建立Stock與Trader對應dict
  3. 新增交易群駔欄位
  4. 交易員分數表
     (1)建立分數表
     (2)計算交易員分數
     (3)格式整理
'''
# =============================================================================

#交易員資料讀取
Trader_Table=pd.read_html('http://172.16.10.13/hedge/makerauthlist.php',encoding='big5')[0]
Trader_Table['群組代碼']=Trader_Table['群組代碼'].astype(str)
Trader_Table.set_index("股票代號",inplace=True)
Trader_Table=Trader_Table.rename(index={'FITX':'99'})

#建立Stock2Trader表
Stock2Trader={}
for stock in set(Trader_Table.index.values):
    if type(Trader_Table.loc[stock,'群組代碼'])==str:
        Stock2Trader[stock]=Trader_Table.loc[stock,'群組代碼']                    
    else:                        
        Stock2Trader[stock]=Trader_Table.loc[stock,'群組代碼'].values[0]

#新增交易群駔欄位
final_data.loc[:,'群組代碼']=final_data.loc[:,'標的代號'].map(lambda v:Stock2Trader.setdefault(v,None))


#交易員分數表(1)建立分數表
TraderScore_Table={}
for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
    table=pd.pivot_table(final_data,index=['群組代碼'],columns=['券商名稱'],values=[scoretype],aggfunc=[np.mean])
    table.columns=table.columns.droplevel(0)
    TraderScore_Table[scoretype]=table
#交易員分數表(2)計算交易員分數
TraderScore_Table_final={}
for key,table in TraderScore_Table.items():
    score=table.rank(pct=True,axis=1,method='max')
    TraderScore_Table_final[key]=score.loc[:,(key,'日盛')].copy()
    
#交易員分數表(3)格式整理  
TraderScore_Table_final=pd.DataFrame(TraderScore_Table_final)
TraderScore_Table_final['平均分數']=TraderScore_Table_final.mean(axis=1)
TraderScore_Table_final=(TraderScore_Table_final*100).round(1)
TraderScore_Table_final['評等']=TraderScore_Table_final['平均分數'].map(table_for_LetterGrade)


del score,TraderScore_Table,table,Trader_Table
del key,scoretype,stock
#%%
# =============================================================================
# 7.標的分數計算
'''
  1. 股票分數表(1)建立分數表
     股票分數表(2)計算股票分數
     股票分數表(3)格式整理 
  2.建立權證分數表
'''
# =============================================================================

#股票分數表(1)建立分數表
StockScore_Table={}
for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
    table=pd.pivot_table(final_data,index=['標的代號'],columns=['券商名稱'],values=[scoretype],aggfunc=[np.mean])
    table.columns=table.columns.droplevel(0)
    StockScore_Table[scoretype]=table
#股票分數表(2)計算股票分數
StockScore_Table_final={}
for key,table in StockScore_Table.items():
    score=table.rank(pct=True,axis=1,method='max')
    StockScore_Table_final[key]=score.loc[:,(key,'日盛')].copy()
#股票分數表(3)格式整理  
StockScore_Table_final=pd.DataFrame(StockScore_Table_final)
StockScore_Table_final['平均分數']=StockScore_Table_final.mean(axis=1)
StockScore_Table_final=(StockScore_Table_final*100).round(1)
StockScore_Table_final=StockScore_Table_final.dropna(how='all')
StockScore_Table_final['評等']=StockScore_Table_final['平均分數'].map(table_for_LetterGrade)
#股票分數表(3)加入群組代號 
Stock2Trader_Series=pd.Series(Stock2Trader, name='交易群組')
StockScore_Table_final=pd.concat([StockScore_Table_final,Stock2Trader_Series],axis=1,join='inner')


#建立權證分數表
Warrant_Score=final_data.loc[:,['標的代號','群組代碼','券商名稱','價格穩定分數','價差大小分數','報價數量分數']]
Warrant_Score=Warrant_Score[Warrant_Score.loc[:,'券商名稱']=='日盛']
Warrant_Score=Warrant_Score.drop('券商名稱',axis=1)
Warrant_Score.loc[:,['價格穩定分數','價差大小分數','報價數量分數']]=(Warrant_Score.loc[:,['價格穩定分數','價差大小分數','報價數量分數']]*100).round(1)
del score,scoretype,table,key,StockScore_Table
del Stock2Trader,Stock2Trader_Series,df,final_data,data_for_BidVolume
#%%
# =============================================================================
# 6.券商
'''
  1.券商累計分數
  2.券商統計檔新增
  3.券商備份軮新增
'''
# =============================================================================
path_broker_statistics=r'T:\權證發行人評等\python版資料\broker_data\券商統計檔.xlsx'
path_broker_backup=r'T:\權證發行人評等\python版資料\broker_data\備份檔\broker.xlsx'
###券商統計檔 讀先前統計檔案
wb= load_workbook(path_broker_statistics)
ws=wb.active
column_max = ws.max_column
row_max = ws.max_row

    ##券商累計分數(openpyxl+pandas)
if row_max>=15:#要有14日平均
    Broker_Statistics={}
    for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
        df=pd.read_excel(path_broker_statistics,scoretype[0:4],index_col='日期')
        Broker_Statistics[scoretype]=df.rolling(15).mean().iloc[-1]
    Broker_Period=pd.DataFrame(Broker_Statistics).round(1)
    Broker_Period['平均分數']=Broker_Period.mean(axis=1).round(1)
    Broker_Period=Broker_Period.sort_values('平均分數',ascending=False)
    Broker_Period['評等']=Broker_Period['平均分數'].map(table_for_LetterGrade)
    print("券商移動平均完成")
day_start_15=str(df.index[-15])
    ##券商統計檔新增(openpyxl)
        #抓已有券商
for row in ws.iter_rows(max_row=1,min_col=2,max_col=column_max, values_only=True):
    brokerlist=list(row) 
        #抓已有資料天數
for col in ws.iter_cols(max_row=row_max,min_row=2,max_col=1,values_only=True):
    datelist=list(col)   
        #確認日期不重複，則新增統計檔
if last_work_day_str in datelist:
    print('已有重複"券商統計檔"')
else:
    for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
        temp=BrokerScore_Table.loc[brokerlist,scoretype].to_list()
        temp.insert(0,last_work_day_str)
        ws=wb[scoretype[0:4]]
        ws.append(temp)
    wb.save(path_broker_statistics)

    ##券商備份新增(openpyxl+pandas)
wb= load_workbook(path_broker_backup)
sheets = wb.sheetnames
        #確認日期不重複，則新增備份
if last_work_day_str in sheets:
    print('已有重複"券商備份檔"')
else:
    writer = pd.ExcelWriter(path_broker_backup, engine = 'openpyxl',mode='a')
    BrokerScore_Table_Sort=BrokerScore_Table.sort_values('總分',ascending=False)
    BrokerScore_Table_Sort.to_excel(writer, sheet_name=str(last_work_day_str))
    writer.save()
BrokerScore_Table=BrokerScore_Table.sort_values('總分',ascending=False)
print('------------------')   
print('券商分數 完成')
print('------------------\n\n')   
del col,column_max,row,row_max,brokerlist
del path_broker_statistics,path_broker_backup,ws,wb,sheets
#%%
# =============================================================================
# 7.交易員
'''
  1.交易員累計分數
  2.交易員計檔新增
  3.交易員備份軮新增
'''
# =============================================================================

path_trader_statistics=r'T:\權證發行人評等\python版資料\trader_data\交易員統計檔.xlsx'
path_trader_backup=r'T:\權證發行人評等\python版資料\trader_data\備份檔\trader.xlsx'
#xw 讀先前統計檔案
app = xw.App(visible=False, add_book=False)
wb = app.books.open(path_trader_statistics);ws=wb.sheets[0]
column_max = ws.range('A1').expand('table').columns.count
row_max = ws.range('A1').expand('table').rows.count

    ##交易員累計分數(openpyxl+pandas)
if row_max>=15:#要有14日平均
    Trader_Statistics={}
    for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
        df=pd.read_excel(path_trader_statistics,scoretype[0:4],index_col='日期')
        Trader_Statistics[scoretype]=df.rolling(15,min_periods=12).mean().iloc[-1]
    Trader_Period=pd.DataFrame(Trader_Statistics).round(1)
    Trader_Period['平均分數']=Trader_Period.mean(axis=1).round(1)
    Trader_Period=Trader_Period.sort_values('平均分數',ascending=False)
    Trader_Period['評等']=Trader_Period['平均分數'].map(table_for_LetterGrade)
    print("交易員移動平均完成")
    
    ##交易員統計檔新增(openpyxl)
            #交易員(統計檔)
traderlist=ws.range('a1').expand('table').value[0][1:]
traderlist=[str(int(trader)) for trader in traderlist]

            #交易員(最新)
wb2=  app.books.open(path_trader_backup)
ws2 = wb2.sheets[-1]
            #是否新增交易員，有的話增加統計檔columns
row_max2 = ws2.range('A1').expand('table').rows.count
traderlist2=ws2.range(f'a1:a{row_max2}').value[1:]

if set(traderlist2)-set(traderlist)==set():
    del traderlist2
else:
    NewTrader=list(set(traderlist2)-set(traderlist))
    traderlist=traderlist.append(NewTrader)
    for n in range(len(wb.sheets)):
        ws=wb.sheets[n-1]
        for i,trader in enumerate(NewTrader):
            ws[0,column_max+i].value=trader

datelist=ws[1:row_max,0].value

    ##交易員備份新增(openpyxl+pandas)
if last_work_day_str in datelist:#日期不重複
    print('已有重複"交易員計檔"')
else:
    for scoretype in ['價格穩定分數','價差大小分數','報價數量分數']:
        temp=TraderScore_Table_final.loc[traderlist,scoretype].to_list()
        temp.insert(0,last_work_day_str)
        ws=wb.sheets[scoretype[0:4]]
        ws.range(row_max+1,1).value=temp
    wb.save(path_trader_statistics)
    
wb.save(path_trader_statistics)
wb.close()
wb2.close()
app.quit()

    ##交易員備份檔新增(openpyxl+pandas)
wb= load_workbook(path_trader_backup)
sheets = wb.sheetnames
if last_work_day_str in sheets:
    print('已有重複"交易員備份檔"')
else:
    writer = pd.ExcelWriter(path_trader_backup, engine = 'openpyxl',mode='a')
    TraderScore_Table_final.to_excel(writer, sheet_name=str(last_work_day_str))
    writer.save()
print('------------------')   
print('交易員分數 完成')
print('------------------\n\n')   
del column_max,row_max,traderlist,datelist,path_trader_statistics,path_trader_backup
del ws,wb,ws2,wb2,sheets,row_max2,app
#%%
# =============================================================================
# 8.股票權證備份檔
'''
  1.股票
  2.權證
'''
# =============================================================================
path=r'T:\權證發行人評等\python版資料\stock_data'
writer = pd.ExcelWriter(path, engine = 'openpyxl',mode='w')
StockScore_Table_final.to_excel(path+'\\'+last_work_day_str+'_stock.xlsx')

path=r'T:\權證發行人評等\python版資料\warrant_data'
writer = pd.ExcelWriter(path, engine = 'openpyxl',mode='w')
Warrant_Score.to_excel(path+'\\'+last_work_day_str+'_warrant.xlsx')
print('------------------')   
print('股票+權證分數 完成')
print('------------------\n\n') 
# #%%
# # =============================================================================
# # Email
# '''
# 1.券商分數表
# 2.交易員分數表
# 3.券商累計
# 4.交易累計
# '''
# # =============================================================================
# df_mail=pd.read_excel(r'Email_發行評等.xls')
# Email_list=df_mail['Email'].dropna().values.tolist()
# Email_str=','.join(Email_list)


# send_user = '新金融商品處<ewarrant@jsun.com>'   #發件人
# receive_users = Email_str   #收件人，可為list
# subject = date_str+'發行人評等'  #郵件主題
# text = """
# 上一交易日評等:
# 券商:
# {table1}
# 交易員:
# {table2}

# 15日平均:
# 券商:
# {table3}
# 交易員:
# {table4}

# 標的分數查詢: T:\權證發行人評等\python版資料\stock_data\{day}_stock.xlsx
# 權證分數查詢: T:\權證發行人評等\python版資料\warrant_data\{day}_warrant.xlsx
# 收件人設定: T:\權證發行人評等\python版資料\Email_發行評等.xls
# """

# html = """
# <html><body>
# <p><font size="5"><b>上一交易日評等:</b></font></p>
# <p>券商分數({day})</p>
# {table1}
# <p>交易員分數({day})</p>
# {table2}
# <hr/>
# <p><font size="5"><b>15交易日平均分數:</b></font></p>
# <p>券商分數({start} ~ {day})</p>
# {table3}
# <p>交易員分數({start} ~ {day})</p>
# {table4}
# <hr/>
# 標的分數查詢: T:\權證發行人評等\python版資料\stock_data\{day}_stock.xlsx<br>
# 權證分數查詢: T:\權證發行人評等\python版資料\warrant_data\{day}_warrant.xlsx<br>
# 收件人設定: T:\權證發行人評等\python版資料\Email_發行評等.xls<br>
# <br>
# 有發現問題或任何建議再通知我<br>
# 昱霖<br>
# </body></html>
# """
# #郵件正文

# server_address = 'mx00.jsun.com'   #伺服器地址

# #構造一個郵件體：正文 附件
# ##-------------一日------------------##
# table_Broker=BrokerScore_Table.to_html()
# table_Broker=re.sub(r'<tr>(.*?券商名稱.*?)</tr>','',table_Broker,flags=re.I|re.M|re.S)
# table_Broker=re.sub(r'<th>.*?','<th bgcolor="#9D9D9D">',table_Broker,flags=re.S)
# table_Broker=re.sub(r'<tr>\n      <th bgcolor="#9D9D9D">日盛</th>'\
#                   ,'<tr bgcolor="#FFAD86">\n      <th bgcolor="#9D9D9D">日盛</th>',table_Broker,flags=re.S)

# table_Trader=TraderScore_Table_final.to_html()
# table_Trader=re.sub(r'<tr>(.*?群組代碼.*?)</tr>','',table_Trader,flags=re.I|re.M|re.S)
# table_Trader=re.sub(r'<th>.*?','<th bgcolor="#9D9D9D">',table_Trader,flags=re.S)

# ##-------------15日------------------##
# table_Broker_Period=Broker_Period.to_html()
# table_Broker_Period=re.sub(r'<th>.*?','<th bgcolor="#9D9D9D">',table_Broker_Period,flags=re.S)    
# table_Broker_Period=re.sub(r'<tr>\n      <th bgcolor="#9D9D9D">日盛</th>'\
#                   ,'<tr bgcolor="#FFAD86">\n      <th bgcolor="#9D9D9D">日盛</th>',table_Broker_Period,flags=re.S)

# table_Trader_Period=Trader_Period.to_html()
# table_Trader_Period=re.sub(r'<th>.*?','<th bgcolor="#9D9D9D">',table_Trader_Period,flags=re.S)    
# table_Trader_Period=re.sub(r'<tr>\n      <th bgcolor="#9D9D9D">日盛</th>'\
#                   ,'<tr bgcolor="#FFAD86">\n      <th bgcolor="#9D9D9D">日盛</th>',table_Trader_Period,flags=re.S)
# ##-------------輸出------------------##
# html = html.format(day=last_work_day_str,start=day_start_15,\
#                    table1=table_Broker,table2=table_Trader,table3=table_Broker_Period,table4=table_Trader_Period)
# text = text.format(table1=tabulate(BrokerScore_Table, headers="firstrow", tablefmt="grid")\
#                   ,table2=tabulate(TraderScore_Table_final, headers="firstrow",tablefmt="grid")\
#                   ,table3=tabulate(Broker_Period, headers="firstrow",tablefmt="grid")\
#                   ,table4=tabulate(Trader_Period, headers="firstrow",tablefmt="grid")\
#                   ,day=last_work_day_str)
    

# #構建正文
# msg = MIMEMultipart("alternative", None,[MIMEText(text), MIMEText(html,'html')])
# #把正文加到郵件體裡面去
# msg['Subject']=subject    #主題
# msg['From']=send_user      #發件人
# msg['To']=receive_users           #收件人
# # 傳送郵件 SMTP
# smtp= SMTP(server_address,25)  # 連線伺服器，SMTP_SSL是安全傳輸

# # #smtp.login(send_user, password)
# smtp.sendmail(send_user, Email_list, msg.as_string())  # 傳送郵件
# print('------------------') 
# print('郵件傳送成功！')
# print('------------------') 


        